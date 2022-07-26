import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
path = r"your file path" #Please make sure this folder contains ONLY your working files, but you can have some subfolders
def get_path(rootdir):
    list_dirs = os.walk(rootdir)
    print(list_dirs) 
    for root, dirs, files in list_dirs:    
        for f in files: 
            yield (os.path.join(root, f))
m=get_path(path)
for i in m:
    if '~' not in i:
        data = load_workbook(i)
        sheet=data['Sheet1']
        maxrow=sheet.max_row
        for n in range(1,maxrow+1):
            cell=str(sheet.cell(n,7).value)
            if cell=='1' or cell=='0.01':
                yellow_fill = PatternFill(fill_type='solid', fgColor="FFFF00")
                sheet.cell(n, 7).fill = yellow_fill
        for a in range(1,maxrow+1):
            cell2=str(sheet.cell(a,9).value)
            if cell2=='N/A':
                yellow_fill = PatternFill(fill_type='solid', fgColor="FFFF00")
                sheet.cell(a, 9).fill = yellow_fill
        data.save(i)
        print(i,'已完成')
print('全部完成')